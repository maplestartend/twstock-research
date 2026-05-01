"""Excel 匯出。

設計取捨：
- 接 list[dict] 而非 Pydantic instance — 避免 app/ 層 import api/schemas（破壞分層），
  且 caller 已從 HoldingRow / RadarHit 用 .model_dump() 拿到 dict，再餵給這裡很自然。
- 回傳 bytes 而非 BytesIO，讓 FastAPI Response 可以直接 `content=bytes`，不用再操作 stream。
- 不在這裡計算任何業務指標 — 純 presentation：欄位排序、儲存格樣式、欄寬、凍結列。
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


# 表頭：brand-500 底白字（與前端 theme 對齊，列印出來辨識度高）
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="FF4F46E5")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _apply_header(ws: Worksheet, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN


def _autosize(ws: Worksheet, headers: list[str], sample_rows: Iterable[list[Any]]) -> None:
    """根據前 50 列估欄寬。中文字符按 1.6 倍寬處理，避免 width=N 在繁中下被截。"""
    samples = list(sample_rows)[:50]
    for i, h in enumerate(headers, start=1):
        max_len = _display_width(h)
        for r in samples:
            if i - 1 < len(r):
                max_len = max(max_len, _display_width(r[i - 1]))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 32)


def _display_width(v: Any) -> int:
    if v is None:
        return 0
    s = str(v)
    width = 0
    for ch in s:
        # CJK Unified Ideographs / Fullwidth：在 Excel 預設字型約佔 1.6 個半形寬，向上取 2
        if "一" <= ch <= "鿿" or "＀" <= ch <= "￯":
            width += 2
        else:
            width += 1
    return width


def _pct(v: Any) -> str | None:
    """0.123 → '12.30%'；None / 非數值 → None。"""
    if v is None:
        return None
    try:
        return f"{float(v) * 100:.2f}%"
    except (TypeError, ValueError):
        return None


def holdings_workbook(rows: list[dict]) -> bytes:
    """持股明細匯出。rows 來自 /api/portfolio/holdings 的 HoldingRow.model_dump()。

    欄位順序模擬使用者在 holdings 頁從左到右掃讀：
    身分 → 部位 → 即時市值 → 損益 → 評分 → 風控 → 備註。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "持股明細"

    headers = [
        "代號", "名稱",
        "張數", "成本", "現價", "今日漲跌",
        "市值", "未實現損益", "未實現 %", "扣費後損益", "扣費後 %",
        "短期", "中期", "長期", "綜合",
        "ATR 停損", "停損距離", "ATR 類型", "進場日",
        "警告",
    ]
    ws.append(headers)
    _apply_header(ws, 1, len(headers))

    sample_data: list[list[Any]] = []
    for r in rows:
        line = [
            r.get("stockId"),
            r.get("stockName"),
            r.get("shares"),
            r.get("avgCost"),
            r.get("price"),
            _pct(r.get("todayPct")),
            r.get("marketValue"),
            r.get("unrealizedPnl"),
            _pct(r.get("unrealizedPnlPct")),
            r.get("netUnrealizedPnl"),
            _pct(r.get("netUnrealizedPnlPct")),
            r.get("shortScore"),
            r.get("midScore"),
            r.get("longScore"),
            r.get("compositeScore"),
            r.get("atrStop"),
            _pct(r.get("atrDistancePct")),
            _atr_kind_label(r.get("atrKind")),
            r.get("entryDate"),
            "; ".join(r.get("warnings") or []),
        ]
        ws.append(line)
        sample_data.append(line)

    ws.freeze_panes = "C2"  # 凍結代號 + 名稱兩欄與表頭，捲動時定位用
    _autosize(ws, headers, sample_data)
    return _to_bytes(wb)


def _atr_kind_label(kind: str | None) -> str | None:
    if kind == "trailing":
        return "追蹤"
    if kind == "fixed":
        return "固定"
    return None


def radar_hits_workbook(
    hits: list[dict],
    *,
    strategy: str,
    market_label: str,
    as_of: str | None,
) -> bytes:
    """雷達命中匯出。hits 來自 /api/radar/scan 的 RadarHit.model_dump()。

    Sheet 標題用策略名 + 市場（不超過 31 字元 Excel 上限），第一列加 metadata
    註記（策略 / 市場 / 截止日 / 命中數），便於使用者寄出後對方理解條件。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = (strategy or "雷達命中")[:31]

    # Metadata row（合併至最後一欄）
    headers = [
        "代號", "名稱", "市場",
        "收盤", "短期", "中期", "長期", "綜合",
        "建議", "VR-MACD", "命中策略",
    ]
    meta_text = (
        f"策略：{strategy or '全部'}　市場：{market_label}　"
        f"截止：{as_of or '-'}　命中：{len(hits)} 檔　"
        f"匯出：{datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )
    ws.append([meta_text])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    meta_cell = ws.cell(row=1, column=1)
    meta_cell.font = Font(italic=True, color="FF6B7280")
    meta_cell.alignment = Alignment(vertical="center")

    ws.append(headers)
    _apply_header(ws, 2, len(headers))

    sample_data: list[list[Any]] = []
    for h in hits:
        line = [
            h.get("stockId"),
            h.get("stockName"),
            h.get("market"),
            h.get("close"),
            h.get("short"),
            h.get("mid"),
            h.get("long"),
            h.get("composite"),
            h.get("recommendation"),
            h.get("vrMacd"),
            h.get("strategies"),
        ]
        ws.append(line)
        sample_data.append(line)

    ws.freeze_panes = "C3"  # 凍結代號名稱 + 表頭 + metadata
    _autosize(ws, headers, sample_data)
    return _to_bytes(wb)


def _to_bytes(wb: Workbook) -> bytes:
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
